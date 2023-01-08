import camelot.io as camelot
import openpyxl

# Read the PDF file into a list of DataFrames
tables = camelot.read_pdf('sample.pdf')

# Initialize the Excel file and add the data
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Iterate over the DataFrames and add them to the Excel file
for i, df in enumerate(tables):
    # Iterate over the rows and columns of the DataFrame and add them to the Excel file
    for j, row in df.iterrows():
        for k, cell in row.iteritems():
            worksheet.cell(row=j+1, column=k+1).value = cell

    # Add a new worksheet for the next DataFrame
    worksheet = workbook.create_sheet()

# Save the Excel file
workbook.save('example.xlsx')
