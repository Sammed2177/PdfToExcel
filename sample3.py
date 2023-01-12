import PyPDF2
from PIL import Image
import openpyxl

# Open the PDF file
with open('demo.pdf', 'rb') as file:
    pdf = PyPDF2.PdfReader(file)

# Extract the text and images from the PDF file
data = []
images = []
for page in pdf.pages:
    # Extract the text from the page
    text = page.extractText()
    data.append(text)

    # Extract the images from the page
    xobjects = page['/Resources']['/XObject'].getObject()
    for obj in xobjects:
        if xobjects[obj]['/Subtype'] == '/Image':
            size = (xobjects[obj]['/Width'], xobjects[obj]['/Height'])
            data = xobjects[obj].getData()
            if xobjects[obj]['/ColorSpace'] == '/DeviceRGB':
                mode = "RGB"
            else:
                mode = "P"
            img = Image.frombytes(mode, size, data)
            images.append(img)

# Create the Excel file and add the data and images to it
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the text to the first column of the worksheet
for row, text in enumerate(data):
    worksheet.cell(row=row+1, column=1).value = text

# Write the images to the remaining columns of the worksheet
for i, image in enumerate(images):
    # Add the image to a new column
    column = i + 2
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 20
    for row, text in enumerate(data):
        worksheet.cell(row=row+1, column=column).value = image

workbook.save('exe.xlsx')
