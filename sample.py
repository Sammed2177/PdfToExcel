import fitz
import requests
import json
import openpyxl

# Open the PDF file using PyMuPDF
pdf = fitz.open('table.pdf')

# Initialize the Excel file and add the data
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Iterate over the pages of the PDF
for i in range(pdf.page_count):
    # Get the page and its dimensions
    page = pdf.load_page(i)
    pw, ph = page.mediabox_size

    # Extract the text from the page
    text = page.get_text("text")

    # Split the text into lines
    lines = text.split('\n')

    # Iterate over the lines and add them to the Excel file
    for j, line in enumerate(lines):
        worksheet.cell(row=j+1, column=1).value = line

    # Extract the images from the page
    images = page.get_image_bbox('JPG')



    # Iterate over the images and add them to the Excel file
    for j, image in enumerate(images):
        # Save the image to a temporary file
        with open('temp.png', 'wb') as f:
            f.write(image[-1])

        # Use the OCR API to recognize the text in the image
        with open('temp.png', 'rb') as f:
            response = requests.post(
                'https://api.ocr.space/parse/image',
                headers={
                    'apikey': 'K84230161688957'
                },
                files={
                    'temp.png': f
                }
            )

        # Extract the recognized text from the API response
        data = json.loads(response.text)
        text = data['ParsedResults'][0]['ParsedText']

        # Split the text into lines
        lines = text.split('\n')

        # Iterate over the lines and add them to the Excel file
        for k, line in enumerate(lines):
            worksheet.cell(row=k+1, column=j+2).value = line

# Save the Excel file
workbook.save('example.xlsx')
